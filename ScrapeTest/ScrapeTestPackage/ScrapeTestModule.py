'''
Created on Apr 13, 2020

@author: niunani
'''
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import re
import urllib
import pandas as pd
import openpyxl
from openpyxl import load_workbook, drawing
from openpyxl.utils import get_column_letter
from datetime import datetime




driver = webdriver.Firefox(executable_path=r'/Users/niunani/Selenium/geckodriver')
driver.get("http://www.petitesannonces.pf/")
#driver.get("http://www.petitesannonces.pf/cherche.php?p=3&c=5&q=Moorea#txt")
#driver.get("http://www.petitesannonces.pf/cherche.php?p=1&c=5&q=Moorea#txt")
#driver.get("http://www.petitesannonces.pf/cherche.php?p=4&c=5&q=Moorea#txt")

def category_chooser(category, locator):
    Select(driver.find_element_by_id(locator)).select_by_visible_text(category)

def search(text_input):
    try:
        search_input = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.ID, 'q'))
            #AC.move_to_element()
            )
        search_input.send_keys(text_input)
        search_input.send_keys(Keys.ENTER)
    finally:
        print('Search finally')
        
def get_children(container_locator, container_by_type, children_by_type, child_locator):
    try:
        children_elements = []
        container_div = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((container_by_type, container_locator)))
        if children_by_type == 'css_selector':
            children_elements = container_div.find_elements_by_css_selector(child_locator)
        elif children_by_type == 'class_name':
            children_elements = container_div.find_elements_by_class_name(child_locator)
        else:
            "No children found"
        return children_elements
    finally:
        print("Finished atttempting to get children of" + container_locator)

def get_item_hrefs():
    try:
        item_list = get_children('che', By.ID, 'class_name', 'lda')
        print(item_list)
        item_href_list = []
        for item in item_list:
            print(item.get_attribute('href'))
            item_href_list.append(item.get_attribute('href'))
        print(item_href_list)
    finally:
        print('loop through finally')
    
    return item_href_list

def get_match_list(regex, text):
    return re.findall(regex, text)

def get_main_text():
    details = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.ID, 'det')))
    return details.text


def check_pro():
    try:
        h3 = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CLASS_NAME, 'pro')))
        print(h3)
        return True
    except:
        #print('Not pro')
        return False
    
def save_image(image_element, file_name):
    src = get_image_source(image_element)
    print(src)
    urllib.urlretrieve(src, file_name)

def get_image_source(image_element):
    return image_element.get_attribute("src")

def get_images(locator):
    image_container = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.ID, locator)))
    images = image_container.find_elements_by_css_selector("img")
    return images

# def get_page_list(locator):
#     try:
#         page_div = WebDriverWait(driver, 5).until(
#             EC.element_to_be_clickable((By.CLASS_NAME, locator)))
#         buttons_list = page_div.find_elements_by_css_selector("a")
#         return buttons_list
#     finally:
#         print("Check and click if last page - completed")
        


def click_next_page(buttons_list):
    try:
#         page_div = WebDriverWait(driver, 5).until(
#             EC.element_to_be_clickable((By.CLASS_NAME, locator)))
#         buttons_list = page_div.find_elements_by_css_selector("a")
#         print(buttons_list)
#         for button in buttons_list:
#             print(button.text)
        #buttons_list = get_children(locator, By.CLASS_NAME, 'a')
        if buttons_list[-1].text.isdigit():
            print("Last page")
        else:
            buttons_list[-1].click()
    finally:
        print("Check and click if last page - completed")

def scrape_posts(driver, search_text, list_dict):
    prix_regex = r'PRIX :.+XPF.+'
    email_regex = r'[\w\.-]+@[\w\.-]+'
    lieux_regex = r'LIEU :.+'
    tel_regex = r'8[79][\d\ ]+'
    src_regex = r'(?=\w+\.\w{3,4}$).+'
    href_list = get_item_hrefs()
    print("hrefs:")
    print(href_list)
    for href in get_item_hrefs():
        driver.get(href)
        #todo: see if there is a way to get around this
        #buttons_list = get_children('pag', By.CLASS_NAME, 'a')
        main_text = get_main_text()
        print(main_text)
        lieu = get_match_list(lieux_regex, main_text)
        print(len(lieu))
        if len(lieu) == 1:
            lieu = get_match_list(lieux_regex, main_text)[0]
            print("All good")
        elif len(lieu) == 0:
            lieu = 'Lieu non-trouve'
        else:
            lieu = get_match_list(lieux_regex, main_text)[0]
            print("Issue with finding 'lieu'")
            print(len(lieu))
           
        print(search_text.lower())
        print(lieu.lower())
        if search_text.lower() in lieu.lower() or lieu == '':
            list_dict[href] = {}
            list_dict[href]["lieu"] = lieu
            list_dict[href]["id"] = get_match_list(r'\d.+',href)[0]
            list_dict[href]["email"] = get_match_list(email_regex, main_text)
            list_dict[href]["prix"] = get_match_list(prix_regex, main_text)
            list_dict[href]["tel"] = get_match_list(tel_regex, main_text)
            images = get_children("pho", By.ID, 'css_selector', 'img')
               
            image_source = get_match_list(src_regex, get_image_source(images[0]))[0]
            print(image_source)
               
            list_dict[href]["img_rel_path"] = image_source
                   
            if check_pro():
                list_dict[href]["type_post"] = "Pro"
            else:
                list_dict[href]["type_post"] = "Particulier"
            for image in images:
                #save_image(image, list_dict[href]["img_src"])
                try:
                    save_image(image, image_source)
                except KeyError:
                    print("No image available")
        else:
            continue
    
    return list_dict

def export_to_excel(result_dict):
    df = pd.DataFrame(result_dict).T
    file_name = str(datetime.now()) + '.xlsx'
    df.to_excel(file_name)
    wb = load_workbook(file_name)
    ws = wb.worksheets[0]
    print('max row' + str(ws.max_row))
    print('max column' + str(ws.max_column))
    column_letter = get_column_letter(ws.max_column + 1)
    
    i = 2
    for index, row in df.iterrows():
        print(index)
        print(row)
        print row['img_rel_path']
        img = openpyxl.drawing.image.Image(row['img_rel_path'])
        img.anchor = column_letter + str(i)
        ws.add_image(img)
        wb.save(file_name)
        i =+ 1
        
    

def run(search_text, category_choice):
#     search_text = 'Moorea'
#     category_chooser('Loue maison', 'c')
    search(search_text)
    list_dict = {}
    page_id_regex = r'p=\d'
    buttons_list = get_children('pag', By.CLASS_NAME, 'css_selector', 'a')
    page_numbers = []
    page_links = []
    print(buttons_list)
    for button in buttons_list:
        page_links.append(button.get_attribute("href"))
        page_numbers.append(int(get_match_list(r'\d+', get_match_list(page_id_regex, button.get_attribute("href"))[0])[0]))
    print(page_numbers)

    i = 0
    (max(page_numbers))
    while i < max(page_numbers) - 1:
        print(page_links[i])
        scrape_posts(driver, search_text, list_dict)
        print("SCRAPING <---------------")
        print(list_dict)
        driver.get(page_links[i])
        i += 1
         
    print(len(list_dict))
    export_to_excel(list_dict)
    driver.close()
    
run('Moorea', category_chooser('Vends terrain', 'c'))

